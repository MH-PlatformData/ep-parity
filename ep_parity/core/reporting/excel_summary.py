"""
Generate Excel summary reports for multi-employer parity comparisons.

This module analyzes comparison reports from multiple employers and creates
comprehensive Excel summaries for leadership review and detailed analysis.

Features:
- Executive Summary: High-level overview for leadership
- Detailed Analysis: File-by-file analysis with review priority, issue type, and comments column
- Color-coded formatting for quick identification of issues
- Trend analysis across employers
- Saves in same directory as comparison report by default
"""

import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ep_parity.utils.logging import get_logger

logger = get_logger(__name__)


class ComparisonReportParser:
    """Parse comparison report text files."""

    def __init__(self, report_path: Path):
        self.report_path = report_path
        self.content = self._read_report()

    def _read_report(self) -> str:
        """Read report file content."""
        try:
            with open(self.report_path, 'r', encoding='utf-8') as f:
                return f.read()
        except Exception as e:
            logger.error(f"Error reading {self.report_path}: {e}")
            return ""

    def parse(self) -> Dict:
        """
        Parse comparison report and extract key metrics.

        Returns:
            Dict with file-level comparison results
        """
        results = {
            'report_file': self.report_path.name,
            'files_compared': 0,
            'files_identical': 0,
            'files_with_differences': 0,
            'total_differences': 0,
            'file_details': []
        }

        if not self.content:
            return results

        # First, extract summary information and create file entries
        self._parse_summary_section(results)
        self._parse_summary_files(results)

        # Then enhance with detailed file sections
        # Look for "File: filename" patterns
        file_pattern = r'File: ([\w\-\.]+)\s*\n'
        file_sections = re.split(file_pattern, self.content)

        # file_sections will be: [before_first_file, filename1, content1, filename2, content2, ...]
        for i in range(1, len(file_sections), 2):
            if i + 1 < len(file_sections):
                filename = file_sections[i]
                content = file_sections[i + 1]
                # Find this file in the summary and enhance it
                for file_info in results['file_details']:
                    if file_info['file_name'] == filename:
                        self._enhance_file_info_from_detail(file_info, content)
                        break

        # Update totals
        results['files_compared'] = len(results['file_details'])
        results['total_differences'] = sum(f['difference_count'] for f in results['file_details'])

        return results

    def _parse_summary_section(self, results: Dict) -> None:
        """Parse the SUMMARY section for quick stats."""
        summary_match = re.search(r'SUMMARY:\s*\n-+\s*\n(.*?)(?:\n\nDETAILED|$)', self.content, re.DOTALL)
        if not summary_match:
            return

        summary_text = summary_match.group(1)

        for line in summary_text.split('\n'):
            if 'MATCH' in line or '\u2705' in line:
                match = re.search(r'([\w\-\.]+\.(?:psv|csv))', line)
                if match:
                    results['files_identical'] += 1
            elif 'DIFF' in line or '\u274c' in line:
                match = re.search(r'([\w\-\.]+\.(?:psv|csv))', line)
                if match:
                    results['files_with_differences'] += 1

    def _parse_summary_files(self, results: Dict) -> None:
        """Parse files from the summary section."""
        summary_match = re.search(r'SUMMARY:\s*\n-+\s*\n(.*?)(?:\n\nDETAILED|$)', self.content, re.DOTALL)
        if not summary_match:
            return

        summary_text = summary_match.group(1)

        for line in summary_text.split('\n'):
            file_info = None

            if 'MATCH' in line or '\u2705' in line:
                match = re.search(r'([\w\-\.]+\.(?:psv|csv))', line)
                if match:
                    file_info = {
                        'file_name': match.group(1),
                        'status': 'IDENTICAL',
                        'difference_count': 0,
                        'row_differences': 0,
                        'value_differences': 0,
                        'primary_only_rows': 0,
                        'replicated_only_rows': 0,
                        'notes': []
                    }
            elif 'DIFF' in line or '\u274c' in line:
                match = re.search(r'([\w\-\.]+\.(?:psv|csv))\s*-\s*(.+)', line)
                if match:
                    filename = match.group(1)
                    description = match.group(2)
                    file_info = {
                        'file_name': filename,
                        'status': 'DIFFERENCES',
                        'difference_count': 1,  # Placeholder
                        'row_differences': 0,
                        'value_differences': 0,
                        'primary_only_rows': 0,
                        'replicated_only_rows': 0,
                        'notes': [description]
                    }

                    # Try to extract row counts from description
                    row_count_match = re.search(r'Primary=(\d+),\s*Replicated=(\d+)', description)
                    if row_count_match:
                        primary = int(row_count_match.group(1))
                        replicated = int(row_count_match.group(2))
                        if primary > replicated:
                            file_info['primary_only_rows'] = primary - replicated
                            file_info['difference_count'] += file_info['primary_only_rows']
                        elif replicated > primary:
                            file_info['replicated_only_rows'] = replicated - primary
                            file_info['difference_count'] += file_info['replicated_only_rows']

            if file_info:
                results['file_details'].append(file_info)
                results['files_compared'] += 1

    def _enhance_file_info_from_detail(self, file_info: Dict, detail_content: str) -> None:
        """Enhance file info with details from the DETAILED RESULTS section."""
        lines = detail_content.strip().split('\n')

        # Extract row counts
        for line in lines:
            if 'Primary rows:' in line:
                match = re.search(r'Primary rows:\s*(\d+)', line)
                if match:
                    file_info['primary_rows'] = int(match.group(1))

            if 'Replicated rows:' in line:
                match = re.search(r'Replicated rows:\s*(\d+)', line)
                if match:
                    file_info['replicated_rows'] = int(match.group(1))

            if 'rows with differences' in line.lower():
                match = re.search(r'(\d+)\s+rows?\s+with\s+differences', line, re.IGNORECASE)
                if match:
                    file_info['row_differences'] = int(match.group(1))
                    if file_info['difference_count'] == 1:  # Placeholder from summary
                        file_info['difference_count'] = file_info['row_differences']

            if 'value differences' in line.lower():
                match = re.search(r'(\d+)\s+value\s+differences', line, re.IGNORECASE)
                if match:
                    file_info['value_differences'] = int(match.group(1))

        # Check for IDENTICAL status
        if 'IDENTICAL' in detail_content or 'No differences found' in detail_content:
            file_info['status'] = 'IDENTICAL'
            file_info['difference_count'] = 0

    def _parse_file_section(self, section: str) -> Dict:
        """Parse individual file comparison section."""
        lines = section.strip().split('\n')

        file_info = {
            'file_name': '',
            'status': 'UNKNOWN',
            'difference_count': 0,
            'row_differences': 0,
            'value_differences': 0,
            'primary_only_rows': 0,
            'replicated_only_rows': 0,
            'notes': []
        }

        # Extract file name
        for line in lines:
            if 'Comparing:' in line or 'File:' in line:
                match = re.search(r'[\w-]+\.(?:psv|csv)', line)
                if match:
                    file_info['file_name'] = match.group(0)
                    break

        # Determine status
        if 'IDENTICAL' in section or 'No differences found' in section:
            file_info['status'] = 'IDENTICAL'
        elif 'DIFFERENCES FOUND' in section or 'differences found' in section:
            file_info['status'] = 'DIFFERENCES'
        elif 'SKIPPED' in section:
            file_info['status'] = 'SKIPPED'
            if 'not found' in section.lower():
                file_info['notes'].append('File not found in one or both databases')
            return file_info

        # Extract difference counts
        for line in lines:
            if 'rows with differences' in line.lower():
                match = re.search(r'(\d+)\s+rows?\s+with\s+differences', line, re.IGNORECASE)
                if match:
                    file_info['row_differences'] = int(match.group(1))
                    file_info['difference_count'] += file_info['row_differences']

            if 'value differences' in line.lower():
                match = re.search(r'(\d+)\s+value\s+differences', line, re.IGNORECASE)
                if match:
                    file_info['value_differences'] = int(match.group(1))

            if 'primary only' in line.lower() or 'only in primary' in line.lower():
                match = re.search(r'(\d+)', line)
                if match:
                    file_info['primary_only_rows'] = int(match.group(1))
                    file_info['difference_count'] += file_info['primary_only_rows']

            if 'replicated only' in line.lower() or 'only in replicated' in line.lower():
                match = re.search(r'(\d+)', line)
                if match:
                    file_info['replicated_only_rows'] = int(match.group(1))
                    file_info['difference_count'] += file_info['replicated_only_rows']

        # Add notes for concerning patterns
        if file_info['primary_only_rows'] > 0:
            file_info['notes'].append(f"{file_info['primary_only_rows']} rows only in primary")
        if file_info['replicated_only_rows'] > 0:
            file_info['notes'].append(f"{file_info['replicated_only_rows']} rows only in replicated")
        if file_info['row_differences'] > 10:
            file_info['notes'].append(f"HIGH: {file_info['row_differences']} rows with differences")

        return file_info


def find_comparison_reports(base_path: Path, emp_ids: List[str], date_str: Optional[str] = None) -> Dict[str, List[Path]]:
    """
    Find comparison reports for specified employers.

    Args:
        base_path: Base path to search
        emp_ids: List of employer IDs
        date_str: Optional date string (e.g., '11-14-25')

    Returns:
        Dict mapping emp_id to list of report paths
    """
    reports = {}

    for emp_id in emp_ids:
        reports[emp_id] = []

        # Search pattern: {emp_id}_{date}_comparison_report.txt
        if date_str:
            pattern = f"{emp_id}_{date_str}_*_comparison_report.txt"
        else:
            pattern = f"{emp_id}_*_comparison_report.txt"

        found = list(base_path.glob(pattern))

        if found:
            # Sort by modification time, most recent first
            found.sort(key=lambda x: x.stat().st_mtime, reverse=True)
            reports[emp_id] = found
            logger.info(f"Found {len(found)} report(s) for employer {emp_id}")
        else:
            logger.warning(f"No reports found for employer {emp_id} with pattern: {pattern}")

    return reports


def create_executive_summary_sheet(wb: Workbook, all_results: Dict[str, Dict]) -> None:
    """Create executive summary worksheet."""
    ws = wb.create_sheet("Summary", 0)

    # Title
    ws['A1'] = "Parity Testing Summary"
    ws['A1'].font = Font(size=16, bold=True)
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].font = Font(size=10, italic=True)

    # Summary statistics
    row = 4
    ws[f'A{row}'] = "Overall Statistics"
    ws[f'A{row}'].font = Font(size=14, bold=True)

    row += 1
    total_employers = len(all_results)
    total_files = sum(r['files_compared'] for r in all_results.values())
    total_identical = sum(r['files_identical'] for r in all_results.values())
    total_differences = sum(r['files_with_differences'] for r in all_results.values())

    stats = [
        ('Total Employers Tested', total_employers),
        ('Total Files Compared', total_files),
        ('Files Identical', total_identical),
        ('Files with Differences', total_differences),
        ('Parity Success Rate', f"{(total_identical/total_files*100):.1f}%" if total_files > 0 else "N/A")
    ]

    for label, value in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

    # Per-employer summary
    row += 2
    ws[f'A{row}'] = "Employer Summary"
    ws[f'A{row}'].font = Font(size=14, bold=True)

    row += 1
    headers = ['Employer ID', 'Files Compared', 'Identical', 'Differences', 'Success Rate', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    row += 1
    for emp_id, results in sorted(all_results.items()):
        files_compared = results['files_compared']
        identical = results['files_identical']
        differences = results['files_with_differences']
        success_rate = (identical / files_compared * 100) if files_compared > 0 else 0

        # Determine status
        if success_rate == 100:
            status = "\u2713 PERFECT"
            fill_color = '70AD47'  # Green
        elif success_rate >= 90:
            status = "\u26a0 MINOR ISSUES"
            fill_color = 'FFC000'  # Yellow
        else:
            status = "\u2717 NEEDS REVIEW"
            fill_color = 'C00000'  # Red

        ws[f'A{row}'] = emp_id
        ws[f'B{row}'] = files_compared
        ws[f'C{row}'] = identical
        ws[f'D{row}'] = differences
        ws[f'E{row}'] = f"{success_rate:.1f}%"
        ws[f'F{row}'] = status

        # Apply color coding
        ws[f'F{row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
        if fill_color == '70AD47':
            ws[f'F{row}'].font = Font(color='FFFFFF', bold=True)

        row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20


def create_detailed_summary_sheet(wb: Workbook, all_results: Dict[str, Dict]) -> None:
    """Create detailed file-by-file analysis worksheet."""
    ws = wb.create_sheet("Detailed Analysis")

    # Headers
    headers = [
        'Review Priority', 'Employer ID', 'File Name', 'Status', 'Issue Type', 'Total Differences',
        'Row Differences', 'Primary Only', 'Replicated Only', 'Notes', 'Review Comments'
    ]

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    row = 2
    for emp_id, results in sorted(all_results.items()):
        for file_detail in results['file_details']:
            # Determine issue type and priority
            if file_detail['status'] == 'IDENTICAL':
                issue_type = 'None'
                priority = 'None'
                fill_color = 'E2EFDA'  # Light green
            elif file_detail['status'] == 'SKIPPED':
                issue_type = 'File Missing'
                priority = 'Check File'
                fill_color = 'D9D9D9'  # Gray
            elif file_detail['primary_only_rows'] > 0 or file_detail['replicated_only_rows'] > 0:
                issue_type = 'Row Count Mismatch'
                priority = 'HIGH'
                fill_color = 'FFE699'  # Light red
            elif file_detail['difference_count'] > 10:
                issue_type = 'Multiple Differences'
                priority = 'MEDIUM'
                fill_color = 'FFF2CC'  # Light yellow
            else:
                issue_type = 'Minor Differences'
                priority = 'LOW'
                fill_color = 'FFFFFF'  # White

            ws[f'A{row}'] = priority
            ws[f'B{row}'] = emp_id
            ws[f'C{row}'] = file_detail['file_name']
            ws[f'D{row}'] = file_detail['status']
            ws[f'E{row}'] = issue_type
            ws[f'F{row}'] = file_detail['difference_count']
            ws[f'G{row}'] = file_detail['row_differences']
            ws[f'H{row}'] = file_detail['primary_only_rows']
            ws[f'I{row}'] = file_detail['replicated_only_rows']
            ws[f'J{row}'] = '; '.join(file_detail['notes']) if file_detail['notes'] else ''
            ws[f'K{row}'] = ''  # Review Comments - empty for user input

            # Apply color to entire row
            for col in range(1, 12):
                ws.cell(row=row, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

            # Bold high priority
            if priority == 'HIGH':
                for col in range(1, 12):
                    ws.cell(row=row, column=col).font = Font(bold=True)

            row += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 50
    ws.column_dimensions['K'].width = 40

    # Add auto-filter
    ws.auto_filter.ref = f"A1:K{row-1}"


def generate_excel_summary(emp_ids: List[str], base_path: Path, output_file: Optional[Path] = None, date_str: Optional[str] = None) -> bool:
    """
    Generate Excel summary workbook.

    Args:
        emp_ids: List of employer IDs
        base_path: Base path to search for reports
        output_file: Output Excel file path (if None, saves in same dir as report)
        date_str: Optional date filter

    Returns:
        True if successful
    """
    logger.info(f"Searching for comparison reports in: {base_path}")

    # Find all comparison reports
    report_paths = find_comparison_reports(base_path, emp_ids, date_str)

    if not any(report_paths.values()):
        logger.error("No comparison reports found for any employer")
        return False

    # Parse all reports
    all_results = {}
    first_report_path = None

    for emp_id, paths in report_paths.items():
        if not paths:
            logger.warning(f"No reports found for employer {emp_id}")
            continue

        # Use most recent report
        report_path = paths[0]
        if first_report_path is None:
            first_report_path = report_path
        logger.info(f"Parsing report for employer {emp_id}: {report_path.name}")

        parser = ComparisonReportParser(report_path)
        results = parser.parse()
        all_results[emp_id] = results

    if not all_results:
        logger.error("No reports could be parsed")
        return False

    # Determine output file location
    if output_file is None:
        # Save in the same directory as the first report
        timestamp = datetime.now().strftime('%Y%m%d_%H%M')
        output_file = first_report_path.parent / f"comparison_summary_{timestamp}.xlsx"

    # Create Excel workbook
    logger.info("Generating Excel summary...")
    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)  # Remove default sheet

    # Create worksheets
    create_executive_summary_sheet(wb, all_results)
    create_detailed_summary_sheet(wb, all_results)

    # Save workbook
    wb.save(output_file)
    logger.info(f"Excel summary saved to: {output_file}")

    return True
